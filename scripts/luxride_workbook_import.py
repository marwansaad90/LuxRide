#!/usr/bin/env python3
"""LuxRide Phase 2 workbook dry-run/import payload builder.

This script is intentionally read-only for the source workbook. It parses the
approved XLSX and reports whether the route matrix is safe to import into the
WordPress booking engine.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("openpyxl is required in the bundled workspace Python runtime") from exc


REQUIRED_HEADERS = {
    "Pickup Location",
    "Destination",
    "نقطة الانطلاق",
    "الوجهة",
    "Sedan One Way (€)",
    "Sedan Round Trip (€)",
    "MPV One Way (€)",
    "MPV Round Trip (€)",
    "Mini Van One Way (€)",
    "Mini Van Round Trip (€)",
    "Trip Name (Return)",
    "اسم الرحلة - عودة",
}
ONE_WAY_TRIP_HEADERS = ("Trip Name (One Way)", "Trip Name (Outbound)")
ONE_WAY_TRIP_AR_HEADERS = ("اسم الرحلة - ذهاب",)
VEHICLES = {
    "sedan": ("Sedan One Way (€)", "Sedan Round Trip (€)"),
    "mpv": ("MPV One Way (€)", "MPV Round Trip (€)"),
    "minivan": ("Mini Van One Way (€)", "Mini Van Round Trip (€)"),
}
PERMIT_DESTINATIONS = {"Cairo", "Luxor", "Aswan", "Sharm El Sheikh"}


def slug(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "route"


def money(value: Any) -> float:
    if not isinstance(value, (int, float)):
        raise ValueError(f"non-numeric price: {value!r}")
    if value < 0:
        raise ValueError(f"negative price: {value!r}")
    return round(float(value), 2)


def classification(return_name: str) -> str:
    return "overnight" if "overnight" in return_name.lower() else "overday"


def recommended_trip_type(return_name: str) -> str:
    return "round_trip" if return_name in {"Overday", "Overnight"} else "one_way"


def detect_table(workbook: Any) -> tuple[str, int, list[tuple[Any, ...]], dict[str, int]]:
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        for row_number, values in enumerate(rows, start=1):
            headers = [str(value).strip() if value is not None else "" for value in values]
            header_set = set(headers)
            if REQUIRED_HEADERS.issubset(header_set) and any(header in header_set for header in ONE_WAY_TRIP_HEADERS):
                return sheet_name, row_number, rows, {header: pos for pos, header in enumerate(headers) if header}
    raise SystemExit("Could not find the LuxRide pricing table header row in the workbook")


def first_existing(index: dict[str, int], headers: tuple[str, ...]) -> str:
    for header in headers:
        if header in index:
            return header
    raise KeyError("none of the expected headers exist: " + ", ".join(headers))


def parse_workbook(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name, header_row, rows, index = detect_table(workbook)
    one_way_trip_header = first_existing(index, ONE_WAY_TRIP_HEADERS)
    one_way_trip_ar_header = first_existing(index, ONE_WAY_TRIP_AR_HEADERS)

    routes: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    seen: dict[tuple[str, str], dict[str, Any]] = {}
    duplicates: list[dict[str, Any]] = []

    for source_row, values in enumerate(rows[header_row:], start=header_row + 1):
        if not any(value not in (None, "") for value in values):
            continue

        pickup = str(values[index["Pickup Location"]] or "").strip()
        destination = str(values[index["Destination"]] or "").strip()
        pickup_ar = str(values[index["نقطة الانطلاق"]] or "").strip()
        destination_ar = str(values[index["الوجهة"]] or "").strip()
        outbound_name = str(values[index[one_way_trip_header]] or "").strip()
        outbound_name_ar = str(values[index[one_way_trip_ar_header]] or "").strip()
        return_name = str(values[index["Trip Name (Return)"]] or "").strip()
        return_name_ar = str(values[index["اسم الرحلة - عودة"]] or "").strip()
        problems: list[str] = []

        if not pickup:
            problems.append("missing pickup")
        if not destination:
            problems.append("missing destination")
        if not outbound_name:
            problems.append("missing outbound trip name")
        if not return_name:
            problems.append("missing return trip name")

        prices: dict[str, dict[str, float]] = {}
        for vehicle, (one_way_header, round_trip_header) in VEHICLES.items():
            try:
                prices[vehicle] = {
                    "one_way": money(values[index[one_way_header]]),
                    "round_trip": money(values[index[round_trip_header]]),
                }
            except ValueError as exc:
                problems.append(f"{vehicle}: {exc}")

        route_code = f"{slug(pickup)}-{slug(destination)}"
        route = {
            "route_code": route_code,
            "source_row": source_row,
            "pickup_key": slug(pickup),
            "pickup_label": pickup,
            "pickup_label_ar": pickup_ar,
            "destination_key": slug(destination),
            "destination_label": destination,
            "destination_label_ar": destination_ar,
            "supported_trip_types": ["one_way", "round_trip"],
            "recommended_trip_type": recommended_trip_type(return_name),
            "round_trip_classification": classification(return_name),
            "airport_fee_applicable": pickup == "Hurghada Airport" or destination == "Hurghada Airport",
            "permit_required": pickup in PERMIT_DESTINATIONS or destination in PERMIT_DESTINATIONS,
            "accommodation_fee_eur": 0,
            "enabled": True,
            "display_order": source_row,
            "outbound_trip_name": outbound_name,
            "outbound_trip_name_ar": outbound_name_ar,
            "return_trip_name": return_name,
            "return_trip_name_ar": return_name_ar,
            "prices": prices,
        }

        if problems:
            invalid.append({"source_row": source_row, "pickup": pickup, "destination": destination, "problems": problems})
            continue

        duplicate_key = (pickup, destination)
        if duplicate_key in seen:
            first = seen[duplicate_key]
            duplicates.append(
                {
                    "pickup": pickup,
                    "destination": destination,
                    "first_source_row": first["source_row"],
                    "source_row": source_row,
                    "identical": first["prices"] == prices,
                    "first_prices": first["prices"],
                    "duplicate_prices": prices,
                }
            )
        else:
            seen[duplicate_key] = route

        routes.append(route)

    duplicate_conflicts = [item for item in duplicates if not item["identical"]]
    summary = {
        "workbook": str(path.resolve()),
        "source_file": path.name,
        "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
        "sheet_names": workbook.sheetnames,
        "parsed_sheet": sheet_name,
        "header_row": header_row,
        "raw_data_rows": len(routes) + len(invalid),
        "valid_rows": len(routes),
        "malformed_rows": len(invalid),
        "duplicate_route_pairs": len(duplicates),
        "duplicate_conflicts": len(duplicate_conflicts),
        "unique_routes_after_first_duplicate_policy": len(seen),
        "price_records_after_first_duplicate_policy": len(seen) * len(VEHICLES),
        "trip_price_values_after_first_duplicate_policy": len(seen) * len(VEHICLES) * 2,
        "round_trip_classifications": Counter(route["round_trip_classification"] for route in seen.values()),
        "invalid": invalid,
        "duplicates": duplicates,
        "clean": not invalid and not duplicate_conflicts,
    }

    return {
        "summary": json.loads(json.dumps(summary, default=dict)),
        "routes": list(seen.values()),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Dry-run LuxRide workbook import")
    parser.add_argument("--workbook", default="LuxRide-Pricelist.xlsx")
    parser.add_argument("--json-out", help="Optional path for normalized import payload JSON")
    parser.add_argument("--strict", action="store_true", help="Return non-zero when dry run is not clean")
    args = parser.parse_args()

    result = parse_workbook(Path(args.workbook))
    print(json.dumps(result["summary"], ensure_ascii=False, indent=2))

    if args.json_out:
        output = Path(args.json_out)
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    return 2 if args.strict and not result["summary"]["clean"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
